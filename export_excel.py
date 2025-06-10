import pandas as pd
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill
import random


def generate_colors(n):
    """Generate n distinct colors for clusters"""
    colors = [
        'FFE6E6', 'E6F3FF', 'E6FFE6', 'FFF0E6', 'F0E6FF',
        'FFFFE6', 'FFE6F0', 'E6FFFF', 'F0FFE6', 'FFE6CC',
        'E6E6FF', 'CCFFE6', 'FFE6B3', 'E6CCFF', 'B3FFE6',
        'FFB3E6', 'B3E6FF', 'E6FFB3', 'FFB3CC', 'CCFFB3',
        'FFD700', 'FFB6C1', '98FB98', 'DDA0DD', 'F0E68C',
        'FFA07A', '20B2AA', 'FFE4B5', 'D3D3D3', 'F5DEB3'
    ]
    
    if n <= len(colors):
        return colors[:n]
    else:
        # Generate additional random colors if needed
        additional_colors = []
        for _ in range(n - len(colors)):
            color = f"{random.randint(200, 255):02X}{random.randint(200, 255):02X}{random.randint(200, 255):02X}"
            additional_colors.append(color)
        return colors + additional_colors


def create_colored_excel(df, cluster_column):
    """Create an Excel file with color-coded clusters"""
    cluster_col = f"{cluster_column}_cluster"
    
    if cluster_col not in df.columns:
        return None
    
    # Sort by cluster to group similar items together
    df_sorted = df.sort_values(by=cluster_col).reset_index(drop=True)
    
    # Get unique clusters and assign colors
    unique_clusters = df_sorted[cluster_col].unique()
    colors = generate_colors(len(unique_clusters))
    cluster_colors = dict(zip(unique_clusters, colors))
    
    # Create Excel file in memory
    output = BytesIO()
    
    # Write to Excel
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Write main data sheet
        df_sorted.to_excel(writer, sheet_name='Clustered_Data', index=False)
        
        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Clustered_Data']
        
        # Apply colors to rows based on clusters
        cluster_col_idx = df_sorted.columns.get_loc(cluster_col) + 1  # +1 for Excel 1-based indexing
        
        for row in range(2, len(df_sorted) + 2):  # Start from row 2 (after header)
            cluster_value = df_sorted.iloc[row-2][cluster_col]
            color_hex = cluster_colors.get(cluster_value, 'FFFFFF')
            fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type='solid')
            
            # Apply color to entire row
            for col in range(1, len(df_sorted.columns) + 1):
                worksheet.cell(row=row, column=col).fill = fill
        
        # Create a summary sheet with cluster information
        cluster_summary = df_sorted.groupby(cluster_col).size().reset_index(name='Count')
        cluster_summary['Color'] = cluster_summary[cluster_col].map(cluster_colors)
        cluster_summary.to_excel(writer, sheet_name='Cluster_Summary', index=False)
        
        # Apply colors to summary sheet
        summary_sheet = writer.sheets['Cluster_Summary']
        for row in range(2, len(cluster_summary) + 2):
            cluster_value = cluster_summary.iloc[row-2][cluster_col]
            color_hex = cluster_colors.get(cluster_value, 'FFFFFF')
            fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type='solid')
            
            for col in range(1, len(cluster_summary.columns) + 1):
                summary_sheet.cell(row=row, column=col).fill = fill
    
    output.seek(0)
    return output.getvalue()